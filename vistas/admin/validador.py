from selenium import webdriver
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import openpyxl
import time

# Carga los datos del archivo Excel
wb = openpyxl.load_workbook('C:/xampp/htdocs/DEWESE/erasmus/vistas/admin/validacion.xlsx')
sheet = wb.active

# Inicia el navegador
driver = webdriver.Chrome()

for row in range(2, sheet.max_row + 1):
    # Abre la página web
    driver.get('http://localhost/dewese/erasmus/index.php?menu=loginSolicitud')

    # Encuentra los campos del formulario e inserta los datos
    dni = driver.find_element(By.NAME, 'dni')
    dni_value = sheet.cell(row=row, column=1).value
    if dni_value is not None:
        dni.send_keys(dni_value)

    localizador = driver.find_element(By.NAME, 'localizador')
    localizador_value = sheet.cell(row=row, column=2).value
    if localizador_value is not None:
        localizador.send_keys(localizador_value)

    contraseña = driver.find_element(By.NAME, 'contraseña')
    contraseña_value = sheet.cell(row=row, column=3).value
    if contraseña_value is not None:
        contraseña.send_keys(contraseña_value)
    # Pulsa el botón de enviar
    submit = driver.find_element(By.CSS_SELECTOR, 'input[type="submit"]')
    submit.click()

    # Espera 3 segundos
    time.sleep(3)


    dni = driver.find_element(By.NAME, 'dni')
    localizador = driver.find_element(By.NAME, 'localizador')
    contraseña = driver.find_element(By.NAME, 'contraseña')

    # Verifica las validaciones y escribe los resultados en el Excel
    if dni.get_attribute("class")=="valido":
        sheet.cell(row=row, column=4).value = "true"
    else:
        sheet.cell(row=row, column=4).value = "false"

    if localizador.get_attribute("class")=="valido":
        sheet.cell(row=row, column=5).value = "true"
    else:
        sheet.cell(row=row, column=5).value = "false"

    if contraseña.get_attribute("class")=="valido":
        sheet.cell(row=row, column=6).value = "true"
    else:
        sheet.cell(row=row, column=6).value = "false"

# Guarda los resultados en el Excel
wb.save('C:/xampp/htdocs/DEWESE/erasmus/vistas/admin/validacion.xlsx')

# Cierra el navegador
driver.quit()

